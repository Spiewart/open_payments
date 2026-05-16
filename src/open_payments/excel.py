"""Excel formatting primitives shared across audit and review workbooks.

These are the small, study-neutral pieces that every reviewer-facing
Excel output benefits from: a consistent color palette, helpers to apply
section coloring + frozen headers + data validation + hyperlinks, and a
workaround for openpyxl's auto-formula-typing on leading-``=`` strings.

The palette intentionally uses neutral names (SOURCE / CMS / EVIDENCE /
REVIEWER) instead of study-specific names (Dean / etc.) so the same
colors can be reused across child apps.
"""
from __future__ import annotations

from typing import Callable, Iterable, Optional, Sequence

import pandas as pd

# --------------------------------------------------------------------------
# Color palette (RGB hex; openpyxl PatternFill compatible)
# --------------------------------------------------------------------------

COLOR_SOURCE = "DEEBF7"          # light blue — source-of-truth (e.g. dean info)
COLOR_CMS = "FFF2CC"             # light yellow — OpenPayments side
COLOR_EVIDENCE = "EDEDED"        # light gray — matcher's filter evidence
COLOR_REVIEWER = "E2EFDA"        # light green — reviewer action zone
COLOR_HEADER_TEXT = "1F1F1F"
COLOR_WARN_RED = "FFC7CE"        # Excel "Bad" cell style — warning red
COLOR_WARN_RED_TEXT = "9C0006"   # dark red text on the warning fill
COLOR_LEGEND_HEADER = "D9D9D9"   # gray for legend section headers

# Backward-compatible aliases for callers that historically used the
# dean-flavored names. New code should prefer the SOURCE-prefixed name.
COLOR_DEAN = COLOR_SOURCE


# --------------------------------------------------------------------------
# Section-styling helper
# --------------------------------------------------------------------------

# A ColumnSpec is (header_name, color_rgb, width). Re-used by section-styling
# helpers to know which color to paint each column with.
ColumnSpec = tuple[str, str, int]


def apply_section_styling(
    ws,
    column_specs: Sequence[ColumnSpec],
    *,
    freeze_header: bool = True,
    body_wrap_text: bool = True,
) -> None:
    """Paint header row + bodies with their section colors, set widths.

    Args:
        ws: openpyxl Worksheet.
        column_specs: list of (header_name, color_rgb, width). Columns not
            found in the sheet are silently skipped, so the caller can
            pass a "superset" spec that covers multiple sheet layouts.
        freeze_header: if True, freeze the header row (panes='A2').
        body_wrap_text: if True, wrap text in body cells.
    """
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    if ws.max_row < 1:
        return

    header_to_idx: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col_idx).value
        if h is not None:
            header_to_idx[str(h)] = col_idx

    header_font = Font(bold=True, color=COLOR_HEADER_TEXT)
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    body_align = Alignment(vertical="top", wrap_text=body_wrap_text)

    fills_by_col: dict[int, PatternFill] = {}
    for header, color, width in column_specs:
        if header not in header_to_idx:
            continue
        col_idx = header_to_idx[header]
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        fills_by_col[col_idx] = PatternFill("solid", fgColor=color)

    for row_idx in range(2, ws.max_row + 1):
        for col_idx, fill in fills_by_col.items():
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = body_align

    if freeze_header:
        ws.freeze_panes = "A2"


def apply_data_validation(
    ws,
    column_header: str,
    allowed_values: Iterable[str],
    *,
    allow_blank: bool = True,
) -> None:
    """Attach a dropdown data-validation to all body cells in a column."""
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    header_to_idx = {
        ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)
    }
    if column_header not in header_to_idx:
        return
    # Header-only sheet (no body rows): nothing to validate. openpyxl's
    # CellRange rejects "C2:C1" with "1 must be greater than 2", so we
    # short-circuit before constructing the range. This case shows up for
    # studies that produce zero NPI matches (e.g. source repos that don't
    # publish NPIs and rely on name-cascade matching).
    if ws.max_row < 2:
        return
    col_idx = header_to_idx[column_header]
    col_letter = get_column_letter(col_idx)
    values = list(allowed_values)
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(values)}"',
        allow_blank=allow_blank,
    )
    dv.error = f"Pick one of: {', '.join(values)}" + (
        " (or leave blank)" if allow_blank else ""
    )
    dv.errorTitle = "Invalid value"
    dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")
    ws.add_data_validation(dv)


def apply_hyperlinks(
    ws,
    column_header: str,
    url_for_value: Callable[[object], Optional[str]],
) -> None:
    """Set per-cell hyperlinks in a column using a value→URL function.

    Cells whose value yields ``None`` from ``url_for_value`` are left
    untouched. Hyperlinked cells get blue underlined text formatting.
    """
    from openpyxl.styles import Font

    header_to_idx = {
        ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)
    }
    if column_header not in header_to_idx:
        return
    col_idx = header_to_idx[column_header]
    link_font = Font(color="0563C1", underline="single")
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        url = url_for_value(cell.value)
        if url:
            cell.hyperlink = url
            cell.font = link_font


def force_string_data_type(ws) -> None:
    """Reclassify any cell with ``data_type='f'`` and a string value back to ``'s'``.

    openpyxl auto-tags cells whose value starts with ``=`` as formulas.
    Section headers like ``=== OVERVIEW ===`` aren't valid formulas, and
    Excel prompts to recover the workbook on open ("Removed Records:
    Formula from /xl/worksheets/sheetN.xml part") if these aren't
    rewritten as strings.

    Call this once per LEGEND sheet (or any sheet with ``=== ===``
    markers) before the workbook is saved.
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == "f" and isinstance(cell.value, str):
                cell.data_type = "s"


def paint_warning_cell(cell, *, fill_rgb: str = COLOR_WARN_RED,
                       text_rgb: str = COLOR_WARN_RED_TEXT) -> None:
    """Paint a single cell warning-red with bold dark-red text.

    Used by the suspicion-coloring helper to flag high-suspicion cells.
    Cell-level scoping by design: callers paint specific cells, not
    whole rows.
    """
    from openpyxl.styles import PatternFill, Font, Alignment

    cell.fill = PatternFill("solid", fgColor=fill_rgb)
    cell.font = Font(bold=True, color=text_rgb)
    cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")


def apply_legend_formatting(
    ws,
    *,
    width_a: int = 35,
    width_b: int = 90,
    section_prefix: str = "===",
) -> None:
    """LEGEND sheet styling: column widths, freeze header, bold section
    headers, and force string-type on any leading-``=`` cells.

    Args:
        ws: openpyxl Worksheet.
        width_a: width for column A (field name).
        width_b: width for column B (explanation).
        section_prefix: rows whose column-A value starts with this prefix
            are styled as section dividers (bold + gray fill).
    """
    from openpyxl.styles import Font, Alignment, PatternFill

    ws.column_dimensions["A"].width = width_a
    ws.column_dimensions["B"].width = width_b
    ws.freeze_panes = "A2"
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor=COLOR_LEGEND_HEADER)
    wrap = Alignment(vertical="top", wrap_text=True)
    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        ws.cell(row=row_idx, column=1).alignment = wrap
        ws.cell(row=row_idx, column=2).alignment = wrap
        if isinstance(val, str) and val.startswith(section_prefix):
            ws.cell(row=row_idx, column=1).font = bold
            ws.cell(row=row_idx, column=1).fill = fill
            ws.cell(row=row_idx, column=2).fill = fill
    force_string_data_type(ws)
